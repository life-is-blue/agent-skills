#!/usr/bin/env python3
"""Convert MPP/MSPDI files to Excel review spreadsheets.

Usage:
    python3 mpp_to_excel.py <file.mpp|file.xml> --output review.xlsx
    python3 mpp_to_excel.py <file.mpp|file.xml>                        # auto-names output
    python3 mpp_to_excel.py <file.mpp|file.xml> --analyze              # 5 sheets with Issues
    python3 mpp_to_excel.py <file.mpp|file.xml> --sheets overview,overdue  # selected sheets

Generates a 5-sheet Excel workbook (6 with --analyze):
  - Overview: project info, workstream progress (Level 1 + Level 2 phases)
  - All Tasks: full task list with color coding (red=overdue, yellow=summary, green=done)
  - Overdue: overdue leaf tasks with days-overdue calculation
  - Workstreams: summary tasks at Level 1-3 with % complete
  - Gantt: week-level bar chart for L1+L2 tasks with progress colors and today line
  - Sub-Gantt: per-workstream blocks with L1→L2→L3 expansion (explicit --sheets sub-gantt)
  - Issues (--analyze only): all diagnostics in one filterable table
"""

import sys
import os
import argparse
from datetime import datetime, timedelta, date

sys.path.insert(0, os.path.dirname(__file__))
from mpp_reader import read_project

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl is required. Install with: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


# Styles
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"), bottom=Side("thin"),
)
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_HEADER = PatternFill("solid", fgColor="C00000")
ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")


def _header_row(ws, row, headers, fill=None):
    """Write a styled header row."""
    fill = fill or HEADER_FILL
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.font = HEADER_FONT
        c.fill = fill
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")


def _cell(ws, row, col, value, fmt=None, fill=None, bold=False, indent=0):
    """Write a styled cell."""
    c = ws.cell(row, col, value)
    c.border = THIN_BORDER
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if bold:
        c.font = Font(bold=True)
    if indent:
        c.alignment = Alignment(indent=indent)
    return c


def _set_widths(ws, widths):
    """Set column widths from a dict."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_overview(wb, proj, tasks, now, now_str):
    """Build the Overview sheet."""
    ws = wb.active
    ws.title = "Overview"

    leaf = [t for t in tasks if not t["summary"]]
    overdue = [t for t in leaf if t.get("finish") and t["finish"] < now_str and t["percent_complete"] < 100]

    info = [
        ("Title", proj.get("title", "")),
        ("Start", proj.get("start", "")[:10]),
        ("Finish", proj.get("finish", "")[:10]),
        ("Total Tasks", len(tasks)),
        ("Summary Tasks", len(tasks) - len(leaf)),
        ("Leaf Tasks", len(leaf)),
        ("Overdue (leaf)", len(overdue)),
        ("Report Date", now.strftime("%Y-%m-%d")),
    ]
    for r, (k, v) in enumerate(info, 1):
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v)

    # Level 1 workstreams
    ws_l1 = [t for t in tasks if t["outline_level"] == 1]
    row = len(info) + 2
    ws.cell(row, 1, "Workstreams (Level 1)").font = Font(bold=True, size=12)
    row += 1
    _header_row(ws, row, ["#", "Workstream", "%", "Start", "Finish"])
    row += 1
    for i, w in enumerate(ws_l1, 1):
        _cell(ws, row, 1, i)
        _cell(ws, row, 2, w["name"])
        _cell(ws, row, 3, w["percent_complete"] / 100, fmt="0%")
        _cell(ws, row, 4, w.get("start", "")[:10])
        _cell(ws, row, 5, w.get("finish", "")[:10])
        row += 1

    # Level 2 phases
    ws_l2 = [t for t in tasks if t["outline_level"] == 2 and t["summary"]]
    if ws_l2:
        row += 1
        ws.cell(row, 1, "Phases (Level 2)").font = Font(bold=True, size=12)
        row += 1
        _header_row(ws, row, ["#", "Phase", "%", "Plan%", "Gap%", "Parent"])
        row += 1
        # Build parent map
        current_parent = ""
        parent_map = {}
        for t in tasks:
            if t["outline_level"] == 1:
                current_parent = t["name"]
            if t["outline_level"] == 2 and t["summary"]:
                parent_map[t["uid"]] = current_parent
        for i, w in enumerate(ws_l2, 1):
            _cell(ws, row, 1, i)
            _cell(ws, row, 2, w["name"])
            _cell(ws, row, 3, w["percent_complete"] / 100, fmt="0%")
            pp = w.get("planned_pct")
            gp = w.get("gap_pct")
            _cell(ws, row, 4, pp / 100 if pp is not None else "", fmt="0%" if pp is not None else None)
            gap_fill = None
            if gp is not None:
                if gp >= 10:
                    gap_fill = RED_FILL
                elif gp >= 5:
                    gap_fill = ORANGE_FILL
                elif gp < 0:
                    gap_fill = GREEN_FILL
            _cell(ws, row, 5, gp / 100 if gp is not None else "", fmt="0%" if gp is not None else None, fill=gap_fill)
            _cell(ws, row, 6, parent_map.get(w["uid"], ""))
            row += 1

    _set_widths(ws, {"A": 20, "B": 55, "C": 12, "D": 10, "E": 10, "F": 20})


def build_all_tasks(wb, tasks, now_str):
    """Build the All Tasks sheet."""
    ws = wb.create_sheet("All Tasks")
    headers = ["UID", "WBS", "Name", "Level", "Summary", "Start", "Finish",
               "% Complete", "Critical", "Milestone", "Duration"]
    _header_row(ws, 1, headers)

    for r, t in enumerate(tasks, 2):
        _cell(ws, r, 1, t["uid"])
        _cell(ws, r, 2, t.get("wbs", ""))
        _cell(ws, r, 3, t["name"], indent=t["outline_level"])
        _cell(ws, r, 4, t["outline_level"])
        _cell(ws, r, 5, "Y" if t["summary"] else "")
        _cell(ws, r, 6, t.get("start", "")[:10])
        _cell(ws, r, 7, t.get("finish", "")[:10])
        _cell(ws, r, 8, t["percent_complete"] / 100, fmt="0%")
        _cell(ws, r, 9, "Y" if t.get("critical") else "")
        _cell(ws, r, 10, "Y" if t.get("milestone") else "")
        _cell(ws, r, 11, t.get("duration", ""))

        # Color coding
        is_overdue = (not t["summary"] and t.get("finish") and
                      t["finish"] < now_str and t["percent_complete"] < 100)
        if is_overdue:
            fill = RED_FILL
        elif t["percent_complete"] >= 100:
            fill = GREEN_FILL
        elif t["summary"]:
            fill = YELLOW_FILL
        else:
            fill = None

        if fill:
            for col in range(1, 12):
                ws.cell(r, col).fill = fill

    _set_widths(ws, {"A": 6, "B": 15, "C": 50, "D": 6, "E": 8,
                     "F": 12, "G": 12, "H": 10, "I": 8, "J": 10, "K": 15})
    ws.auto_filter.ref = f"A1:K{len(tasks) + 1}"


def build_overdue(wb, tasks, now, now_str):
    """Build the Overdue sheet."""
    ws = wb.create_sheet("Overdue")
    leaf = [t for t in tasks if not t["summary"]]
    overdue = [t for t in leaf
               if t.get("finish") and t["finish"] < now_str and t["percent_complete"] < 100]

    _header_row(ws, 1, ["UID", "WBS", "Task", "Due Date", "% Complete", "Days Overdue"],
                fill=RED_HEADER)

    for r, t in enumerate(sorted(overdue, key=lambda x: x.get("finish", "")), 2):
        _cell(ws, r, 1, t["uid"])
        _cell(ws, r, 2, t.get("wbs", ""))
        _cell(ws, r, 3, t["name"])
        _cell(ws, r, 4, t.get("finish", "")[:10])
        _cell(ws, r, 5, t["percent_complete"] / 100, fmt="0%")
        try:
            due = datetime.fromisoformat(t["finish"].split("+")[0].replace("Z", ""))
            days = (now - due).days
        except (ValueError, KeyError):
            days = "?"
        _cell(ws, r, 6, days)
        for col in range(1, 7):
            ws.cell(r, col).fill = RED_FILL

    _set_widths(ws, {"A": 6, "B": 15, "C": 50, "D": 12, "E": 10, "F": 14})


def build_issues(wb, issues_list):
    """Build the Issues sheet from analyze_project() output."""
    ws = wb.create_sheet("Issues")
    headers = ["Severity", "Category", "WBS", "Task", "Issue", "Detail"]
    _header_row(ws, 1, headers)

    sev_fills = {
        "CRITICAL": RED_FILL,
        "HIGH": ORANGE_FILL,
        "MEDIUM": YELLOW_FILL,
    }

    for r, issue in enumerate(issues_list, 2):
        fill = sev_fills.get(issue["severity"])
        _cell(ws, r, 1, issue["severity"], fill=fill, bold=True)
        _cell(ws, r, 2, issue["category"])
        _cell(ws, r, 3, issue["wbs"])
        _cell(ws, r, 4, issue["name"])
        _cell(ws, r, 5, issue["issue"])
        _cell(ws, r, 6, issue["detail"])

    _set_widths(ws, {"A": 12, "B": 14, "C": 14, "D": 50, "E": 20, "F": 40})
    if issues_list:
        ws.auto_filter.ref = f"A1:F{len(issues_list) + 1}"
    ws.freeze_panes = "A2"


def build_workstreams(wb, tasks):
    """Build the Workstreams sheet (dynamic depth)."""
    ws = wb.create_sheet("Workstreams")
    # Dynamic: include all summary tasks regardless of level
    ws_all = [t for t in tasks if t["summary"] and t["outline_level"] >= 1]

    _header_row(ws, 1, ["Level", "WBS", "Workstream", "% Complete", "Start", "Finish", "Critical"])

    for r, t in enumerate(ws_all, 2):
        _cell(ws, r, 1, t["outline_level"])
        _cell(ws, r, 2, t.get("wbs", ""))
        _cell(ws, r, 3, t["name"], indent=t["outline_level"] - 1)
        _cell(ws, r, 4, t["percent_complete"] / 100, fmt="0%")
        _cell(ws, r, 5, t.get("start", "")[:10])
        _cell(ws, r, 6, t.get("finish", "")[:10])
        _cell(ws, r, 7, "Y" if t.get("critical") else "")
        if t["outline_level"] <= 2:
            for col in range(1, 8):
                ws.cell(r, col).font = Font(bold=True)
        if t["outline_level"] == 1:
            for col in range(1, 8):
                ws.cell(r, col).fill = YELLOW_FILL

    _set_widths(ws, {"A": 6, "B": 15, "C": 50, "D": 12, "E": 12, "F": 12, "G": 8})


# ---------------------------------------------------------------------------
# Gantt chart
# ---------------------------------------------------------------------------

# Gantt bar fills
GANTT_DONE = PatternFill("solid", fgColor="70AD47")      # green — completed portion
GANTT_PROGRESS = PatternFill("solid", fgColor="70AD47")   # green — done portion (unified with DONE)
GANTT_REMAIN = PatternFill("solid", fgColor="D9D9D9")     # gray — remaining / not started
GANTT_OVERDUE = PatternFill("solid", fgColor="FF4B4B")    # red — overdue
GANTT_MILESTONE = PatternFill("solid", fgColor="FFC000")  # gold — milestone marker
GANTT_L1_FILL = PatternFill("solid", fgColor="D6E4F0")    # light blue — L1 row info cols
GANTT_TODAY = PatternFill("solid", fgColor="4472C4")       # blue — today marker
TODAY_BORDER = Border(left=Side("thick", color="4472C4"))  # blue left border for today line

GANTT_HEADER_MONTH = PatternFill("solid", fgColor="2F5496")
GANTT_HEADER_WEEK = PatternFill("solid", fgColor="4472C4")


def _parse_date(s):
    """Parse ISO date string to date object. Returns None on failure."""
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.split("+")[0].replace("Z", "")).date()
    except (ValueError, TypeError):
        return None


def _monday_of(d):
    """Return Monday of the week containing date d."""
    return d - timedelta(days=d.weekday())


# ---------------------------------------------------------------------------
# Shared Gantt drawing helpers
# ---------------------------------------------------------------------------

def _draw_month_headers(ws, start_row, weeks, info_cols):
    """Draw month header row at start_row. Merges cells across each month span."""
    row = start_row
    cur_month = None
    month_start_col = None
    for i, w in enumerate(weeks):
        col = info_cols + 1 + i
        month_key = (w.year, w.month)
        if month_key != cur_month:
            if cur_month is not None and month_start_col is not None:
                prev_label = f"{datetime(cur_month[0], cur_month[1], 1).strftime('%b')} {cur_month[0]}"
                for mc in range(month_start_col, col):
                    c = ws.cell(row, mc)
                    c.fill = GANTT_HEADER_MONTH
                    c.font = HEADER_FONT
                    c.alignment = Alignment(horizontal="center")
                ws.cell(row, month_start_col, prev_label)
                if col - 1 > month_start_col:
                    ws.merge_cells(start_row=row, start_column=month_start_col,
                                   end_row=row, end_column=col - 1)
            cur_month = month_key
            month_start_col = col
    # Fill + merge last month
    if month_start_col is not None:
        last_col = info_cols + len(weeks)
        for mc in range(month_start_col, last_col + 1):
            c = ws.cell(row, mc)
            c.fill = GANTT_HEADER_MONTH
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center")
        ws.cell(row, month_start_col,
                f"{weeks[-1].strftime('%b')} {weeks[-1].year}")
        if last_col > month_start_col:
            ws.merge_cells(start_row=row, start_column=month_start_col,
                           end_row=row, end_column=last_col)


def _draw_week_headers(ws, start_row, weeks, info_cols, info_labels=None):
    """Draw week date header row.

    info_labels defaults to ["", "Phase", "Actual%", "Plan%", "Gap%", "Start", "Finish"].
    """
    if info_labels is None:
        info_labels = ["", "Phase", "Actual%", "Plan%", "Gap%", "Start", "Finish"]
    _header_row(ws, start_row, info_labels)
    for i, w in enumerate(weeks):
        col = info_cols + 1 + i
        c = ws.cell(start_row, col, w.strftime("%m-%d"))
        c.font = Font(size=8, color="FFFFFF")
        c.fill = GANTT_HEADER_WEEK
        c.alignment = Alignment(horizontal="center", text_rotation=90)
        c.border = THIN_BORDER


def _draw_gantt_row(ws, row, t, weeks, info_cols, today, is_l1=False, bold_l2=False):
    """Draw one task row: info columns (name, %, dates) + gantt bar cells.

    is_l1  – True for L1 summary rows (bold, GANTT_L1_FILL, milestone marker ▼).
    bold_l2 – True to make L2 rows bold (used by Sub-Gantt; default False keeps
              build_gantt behaviour unchanged).

    Indentation is derived from outline_level:
      L1 (is_l1=True): indent 0
      L2 (ol=2):       indent 1
      L3 (ol=3):       indent 2
    """
    ol = t["outline_level"]
    pct = t["percent_complete"]
    t_start = _parse_date(t.get("start", ""))
    t_finish = _parse_date(t.get("finish", ""))

    info_fill = GANTT_L1_FILL if is_l1 else None
    is_bold = is_l1 or (bold_l2 and ol == 2)
    indent = 0 if is_l1 else (ol - 1)

    _cell(ws, row, 1, "■" if is_l1 else "", bold=is_bold, fill=info_fill)
    _cell(ws, row, 2, t["name"], bold=is_bold, indent=indent, fill=info_fill)
    _cell(ws, row, 3, pct / 100, fmt="0%", bold=is_bold, fill=info_fill)

    pp = t.get("planned_pct")
    gp = t.get("gap_pct")
    _cell(ws, row, 4, pp / 100 if pp is not None else "",
          fmt="0%" if pp is not None else None, fill=info_fill)
    gap_fill = info_fill
    if gp is not None:
        if gp >= 10:
            gap_fill = RED_FILL
        elif gp >= 5:
            gap_fill = ORANGE_FILL
        elif gp < 0:
            gap_fill = GREEN_FILL
    _cell(ws, row, 5, gp / 100 if gp is not None else "",
          fmt="0%" if gp is not None else None, fill=gap_fill)
    _cell(ws, row, 6, t.get("start", "")[:10], fill=info_fill)
    _cell(ws, row, 7, t.get("finish", "")[:10], fill=info_fill)

    ws.row_dimensions[row].height = 22

    if not t_start or not t_finish:
        return

    # Calculate progress split point
    total_days = max((t_finish - t_start).days, 1)
    done_days = total_days * pct / 100.0
    progress_date = t_start + timedelta(days=done_days)
    is_overdue = t_finish < today and pct < 100

    # Fill week cells
    for i, w_start in enumerate(weeks):
        col = info_cols + 1 + i
        w_end = w_start + timedelta(days=6)

        if w_end < t_start or w_start > t_finish:
            continue  # outside range

        if pct >= 100:
            fill = GANTT_DONE
        elif pct == 0:
            fill = GANTT_OVERDUE if is_overdue else GANTT_REMAIN
        else:
            week_mid = w_start + timedelta(days=3)
            if week_mid <= progress_date:
                fill = GANTT_PROGRESS
            else:
                fill = GANTT_OVERDUE if is_overdue else GANTT_REMAIN

        c = ws.cell(row, col)
        c.fill = fill
        c.border = THIN_BORDER

        # Milestone marker: L1 finish week only
        if is_l1 and t_finish >= w_start and t_finish <= w_end:
            c.value = "▼"
            c.font = Font(size=9, bold=True, color="000000")
            c.alignment = Alignment(horizontal="center", vertical="center")


def _draw_today_line(ws, start_row, end_row, weeks, info_cols, today):
    """Draw blue today line (thick left border) across rows start_row..end_row."""
    today_col = None
    for i, w_start in enumerate(weeks):
        if w_start <= today <= w_start + timedelta(days=6):
            today_col = info_cols + 1 + i
            break

    if today_col:
        for r in range(start_row, end_row + 1):
            c = ws.cell(r, today_col)
            c.border = Border(
                left=Side("thick", color="4472C4"),
                right=c.border.right if c.border else Side("thin"),
                top=c.border.top if c.border else Side("thin"),
                bottom=c.border.bottom if c.border else Side("thin"),
            )


def _draw_legend(ws, start_row):
    """Draw legend row at start_row."""
    legend_items = [
        (GANTT_DONE, "Completed"),
        (GANTT_REMAIN, "Remaining"),
        (GANTT_OVERDUE, "Overdue"),
        (GANTT_MILESTONE, "Milestone"),
    ]
    ws.cell(start_row, 2, "Legend:").font = Font(size=9, bold=True)
    col_offset = 3
    for fill, label in legend_items:
        c = ws.cell(start_row, col_offset)
        c.fill = fill
        c.border = THIN_BORDER
        ws.cell(start_row, col_offset + 1, label).font = Font(size=8)
        col_offset += 3
    # Today line marker in legend
    c = ws.cell(start_row, col_offset)
    c.border = Border(left=Side("thick", color="4472C4"))
    ws.cell(start_row, col_offset + 1, "Today").font = Font(size=8, color="4472C4", bold=True)


# ---------------------------------------------------------------------------
# build_gantt  (refactored to use shared helpers — behaviour unchanged)
# ---------------------------------------------------------------------------

def build_gantt(wb, tasks, now):
    """Build a Gantt chart sheet with weekly columns for L1+L2 tasks."""
    ws = wb.create_sheet("Gantt")
    today = now.date() if isinstance(now, datetime) else now

    # Collect L1 + L2 summary tasks
    gantt_tasks = []
    for t in tasks:
        ol = t["outline_level"]
        if ol == 1:
            gantt_tasks.append(t)
        elif ol == 2 and t["summary"]:
            gantt_tasks.append(t)

    if not gantt_tasks:
        return

    # Determine project date range
    all_starts = [_parse_date(t.get("start", "")) for t in gantt_tasks]
    all_finishes = [_parse_date(t.get("finish", "")) for t in gantt_tasks]
    all_starts = [d for d in all_starts if d]
    all_finishes = [d for d in all_finishes if d]

    if not all_starts or not all_finishes:
        return

    proj_start = _monday_of(min(all_starts))
    proj_end = max(all_finishes)
    # Extend to end of that week
    proj_end = proj_end + timedelta(days=(6 - proj_end.weekday()))

    # Build week columns
    weeks = []
    w = proj_start
    while w <= proj_end:
        weeks.append(w)
        w += timedelta(days=7)

    info_cols = 7  # A=Level, B=Phase, C=Actual%, D=Plan%, E=Gap%, F=Start, G=Finish
    from openpyxl.utils import get_column_letter

    # Row 1: Month headers
    _draw_month_headers(ws, 1, weeks, info_cols)

    # Row 2: Week date headers
    _draw_week_headers(ws, 2, weeks, info_cols)

    # Data rows
    for r_idx, t in enumerate(gantt_tasks):
        row = 3 + r_idx
        is_l1 = (t["outline_level"] == 1)
        _draw_gantt_row(ws, row, t, weeks, info_cols, today, is_l1=is_l1)

    # Today line (covers header rows + all data rows)
    _draw_today_line(ws, 1, 2 + len(gantt_tasks), weeks, info_cols, today)

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 7
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    for i in range(len(weeks)):
        col_letter = get_column_letter(info_cols + 1 + i)
        ws.column_dimensions[col_letter].width = 3.5

    # Freeze panes: freeze info columns + header rows
    ws.freeze_panes = f"{get_column_letter(info_cols + 1)}3"

    # Legend row at bottom
    legend_row = 3 + len(gantt_tasks) + 2  # skip one blank row
    _draw_legend(ws, legend_row)


# ---------------------------------------------------------------------------
# build_sub_gantt  (per-workstream blocks with L1→L2→L3 expansion)
# ---------------------------------------------------------------------------

def build_sub_gantt(wb, tasks, now):
    """Build Sub-Gantt sheet: one block per L1 workstream, expanded to L3.

    Each block has its own complete header (title row + month row + week row),
    making it easy to screenshot a single workstream for PPT.
    Sub-Gantt is NOT included in --sheets all; it must be requested explicitly.
    """
    ws = wb.create_sheet("Sub-Gantt")
    today = now.date() if isinstance(now, datetime) else now

    # 1. Group tasks into workstreams: L1 → [L2/L3 children]
    workstreams = []
    current_ws_data = None
    for t in tasks:
        ol = t["outline_level"]
        if ol == 1:
            current_ws_data = {"l1": t, "children": []}
            workstreams.append(current_ws_data)
        elif current_ws_data is not None and ol in (2, 3):
            current_ws_data["children"].append(t)

    if not workstreams:
        return

    # 2. Global time range (all workstreams share same columns for alignment)
    all_tasks_flat = []
    for ws_data in workstreams:
        all_tasks_flat.append(ws_data["l1"])
        all_tasks_flat.extend(ws_data["children"])

    all_starts = [_parse_date(t.get("start", "")) for t in all_tasks_flat]
    all_finishes = [_parse_date(t.get("finish", "")) for t in all_tasks_flat]
    all_starts = [d for d in all_starts if d]
    all_finishes = [d for d in all_finishes if d]

    if not all_starts or not all_finishes:
        return

    proj_start = _monday_of(min(all_starts))
    proj_end = max(all_finishes) + timedelta(days=(6 - max(all_finishes).weekday()))

    weeks = []
    w = proj_start
    while w <= proj_end:
        weeks.append(w)
        w += timedelta(days=7)

    info_cols = 7
    from openpyxl.utils import get_column_letter

    total_cols = info_cols + len(weeks)

    # 3. Draw each workstream block
    current_row = 1
    for ws_data in workstreams:
        l1 = ws_data["l1"]
        children = ws_data["children"]

        block_start = current_row

        # Workstream title row (full-width merge, deep blue)
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=total_cols)
        title_cell = ws.cell(current_row, 1, l1["name"])
        title_cell.font = Font(size=12, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="2F5496")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 28
        current_row += 1

        # Month header row
        _draw_month_headers(ws, current_row, weeks, info_cols)
        current_row += 1

        # Week header row (label "Task" instead of "Phase")
        _draw_week_headers(ws, current_row, weeks, info_cols,
                           info_labels=["", "Task", "Actual%", "Plan%", "Gap%", "Start", "Finish"])
        current_row += 1

        # L1 summary row
        _draw_gantt_row(ws, current_row, l1, weeks, info_cols, today, is_l1=True)
        current_row += 1

        # L2 and L3 rows
        for child in children:
            ol = child["outline_level"]
            _draw_gantt_row(ws, current_row, child, weeks, info_cols, today,
                            is_l1=False, bold_l2=(ol == 2))
            current_row += 1

        block_end = current_row - 1

        # Today line over entire block
        _draw_today_line(ws, block_start, block_end, weeks, info_cols, today)

        # Two blank rows between blocks
        current_row += 2

    # 4. Single legend at bottom
    _draw_legend(ws, current_row)

    # 5. Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 50   # wider for L3 task names
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 7
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    for i in range(len(weeks)):
        col_letter = get_column_letter(info_cols + 1 + i)
        ws.column_dimensions[col_letter].width = 3.5

    # No freeze_panes: repeated headers make freezing meaningless


def main():
    parser = argparse.ArgumentParser(description="Convert MPP/MSPDI to Excel review spreadsheet")
    parser.add_argument("file", help="Input file (.mpp or .xml)")
    parser.add_argument("--output", "-o", help="Output .xlsx path (default: auto-named)")
    parser.add_argument("--sheets", default="all",
                        help="Comma-separated: overview,tasks,overdue,workstreams,gantt,sub-gantt or "
                             "'all' (sub-gantt requires explicit inclusion, e.g. --sheets all,sub-gantt)")
    parser.add_argument("--analyze", action="store_true",
                        help="Include Issues sheet with diagnostic analysis")
    args = parser.parse_args()

    if not os.path.isfile(args.file):
        print(f"Error: File not found: {args.file}", file=sys.stderr)
        sys.exit(2)

    # Determine output path
    if args.output:
        output = args.output
    else:
        base = os.path.splitext(os.path.basename(args.file))[0]
        # Sanitize for filesystem
        safe_base = base.replace(":", "-").replace("/", "-")
        output = os.path.join(os.path.dirname(args.file), f"{safe_base}.xlsx")

    # Parse sheets selection
    # "all" alone → standard 5 sheets (sub-gantt excluded by default)
    # "all,sub-gantt" → standard 5 + sub-gantt
    # explicit list → exactly those sheets
    raw_sheets_input = args.sheets.strip()
    if raw_sheets_input == "all":
        sheets = {"overview", "tasks", "overdue", "workstreams", "gantt"}
        if args.analyze:
            sheets.add("issues")
    else:
        parts = {s.strip().lower() for s in raw_sheets_input.split(",")}
        if "all" in parts:
            sheets = {"overview", "tasks", "overdue", "workstreams", "gantt"}
            if args.analyze:
                sheets.add("issues")
            # Add any extras beyond "all"
            sheets |= parts - {"all"}
        else:
            sheets = parts

    # Read project data
    try:
        data = read_project(args.file)
    except Exception as e:
        print(f"Error reading file: {e}", file=sys.stderr)
        sys.exit(1)

    proj = data["project"]
    tasks = data["tasks"]
    now = datetime.now()
    now_str = now.isoformat()

    wb = Workbook()

    if "overview" in sheets:
        build_overview(wb, proj, tasks, now, now_str)
    else:
        wb.remove(wb.active)

    if "tasks" in sheets:
        build_all_tasks(wb, tasks, now_str)

    if "overdue" in sheets:
        build_overdue(wb, tasks, now, now_str)

    if "workstreams" in sheets:
        build_workstreams(wb, tasks)

    if "gantt" in sheets:
        build_gantt(wb, tasks, now)

    if "sub-gantt" in sheets:
        build_sub_gantt(wb, tasks, now)

    if "issues" in sheets:
        from mpp_analyze import analyze_project, build_issues_list
        analysis = analyze_project(data, today=now_str)
        issues_list = build_issues_list(analysis, today=now_str)
        build_issues(wb, issues_list)

    # Ensure output directory exists
    output_dir = os.path.dirname(os.path.abspath(output))
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output)

    # Summary
    leaf = [t for t in tasks if not t["summary"]]
    overdue = [t for t in leaf
               if t.get("finish") and t["finish"] < now_str and t["percent_complete"] < 100]
    print(f"Saved: {output}")
    print(f"Title: {proj.get('title', '')}")
    print(f"Tasks: {len(tasks)} (leaf: {len(leaf)}, overdue: {len(overdue)})")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
