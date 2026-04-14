#!/usr/bin/env python3
"""Plan vs Actual calculator for MPP/MSPDI project files.

Generates a Plan vs Actual gap table matching the Board Meeting format.
Supports multiple MPP files, multi-date forecasting, and Excel output.

Usage:
    # Single file, single date
    python3 mpp_plan_vs_actual.py <file.mpp>

    # Multiple files merged
    python3 mpp_plan_vs_actual.py file1.mpp file2.mpp

    # Weekly forecast (this week, next week, W+2)
    python3 mpp_plan_vs_actual.py <file.mpp> --weeks 3

    # Specific dates
    python3 mpp_plan_vs_actual.py <file.mpp> --dates 2026-04-18,2026-04-25,2026-05-02

    # Excel output
    python3 mpp_plan_vs_actual.py <file.mpp> --weeks 3 --excel gap-report.xlsx

    # JSON output
    python3 mpp_plan_vs_actual.py <file.mpp> --json

Methodology:
    - Milestone Task: leaf task count per workstream (outline level 1)
    - Plan: leaf tasks with Finish <= cutoff date
    - Actual: duration-weighted sum (sum of duration * percent_complete / 100)
    - Target%: Plan / Milestone Task
    - Actual%: Actual / Milestone Task
    - Gap: Target% - Actual%
"""

import sys
import os
import argparse
import json
from datetime import date, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))


# ---------------------------------------------------------------------------
# MPP reading
# ---------------------------------------------------------------------------

def _dur_hours(task):
    """Extract duration in hours from MPXJ task object."""
    d = task.getDuration()
    if d is None:
        return 0.0
    val = d.getDuration()
    return float(val) if val else 0.0


def _read_mpp_tasks(file_path):
    """Read MPP and return flat list of task dicts with all needed fields."""
    from helpers.jpype_bootstrap import get_reader

    reader = get_reader()
    proj = reader.read(file_path)
    title = str(proj.getProjectProperties().getProjectTitle() or
                os.path.basename(file_path))

    tasks = []
    for t in proj.getTasks():
        if t.getName() is None:
            continue
        finish = t.getFinish()
        bl_finish = t.getBaselineFinish()
        tasks.append({
            "name": str(t.getName()),
            "outline_level": int(t.getOutlineLevel() or 0),
            "summary": bool(t.getSummary()),
            "finish": str(finish) if finish else "",
            "baseline_finish": str(bl_finish) if bl_finish else "",
            "finish_variance": str(t.getFinishVariance() or ""),
            "percent_complete": float(t.getPercentageComplete() or 0),
            "duration_hours": _dur_hours(t),
            "wbs": str(t.getWBS() or ""),
            "planned_pct": float(str(t.getNumber(4) or 0)) if t.getNumber(4) else None,
            "gap_pct": float(str(t.getNumber(3) or 0)) if t.getNumber(3) else None,
        })

    return title, tasks


def _read_xml_tasks(file_path):
    """Read MSPDI XML and return flat list of task dicts."""
    from helpers.mspdi_ns import findall, findtext
    from helpers.duration_utils import pt_to_hours
    import xml.etree.ElementTree as ET

    tree = ET.parse(file_path)
    root = tree.getroot()
    title = findtext(root, "Title", os.path.basename(file_path))

    tasks = []
    for task_el in findall(root, "Tasks/Task"):
        name = findtext(task_el, "Name", "")
        if not name:
            continue
        tasks.append({
            "name": name,
            "outline_level": int(findtext(task_el, "OutlineLevel", "0")),
            "summary": findtext(task_el, "Summary", "0") == "1",
            "finish": findtext(task_el, "Finish", ""),
            "baseline_finish": findtext(task_el, "BaselineFinish", ""),
            "finish_variance": "",
            "percent_complete": float(findtext(task_el, "PercentComplete", "0")),
            "duration_hours": pt_to_hours(findtext(task_el, "Duration", "PT0H0M0S")),
            "wbs": findtext(task_el, "WBS", ""),
            "planned_pct": None,
            "gap_pct": None,
        })

    return title, tasks


def read_tasks(file_path):
    """Read MPP or XML file and return (title, tasks)."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xml":
        return _read_xml_tasks(file_path)
    return _read_mpp_tasks(file_path)


# ---------------------------------------------------------------------------
# Core calculation
# ---------------------------------------------------------------------------

def _build_workstreams(all_tasks):
    """Group tasks by Level 1 workstream. Returns (ws_order, ws_tasks)."""
    ws_order = []
    ws_tasks = {}
    current_ws = None

    for t in all_tasks:
        ol = t["outline_level"]
        if ol == 1:
            current_ws = t["name"]
            ws_order.append(current_ws)
            ws_tasks[current_ws] = {
                "meta": t,
                "leaves": [],
            }
        elif ol > 1 and not t["summary"] and current_ws is not None:
            ws_tasks[current_ws]["leaves"].append(t)

    return ws_order, ws_tasks


def _calc_ws(leaves, cutoff_str):
    """Calculate Plan vs Actual for a list of leaf tasks at a given cutoff."""
    total = len(leaves)
    if total == 0:
        return {"total": 0, "plan": 0, "actual": 0.0,
                "target_pct": 0, "actual_pct": 0, "gap": 0}

    plan = 0
    actual_w = 0.0
    total_dur = 0.0

    for t in leaves:
        dur = t["duration_hours"]
        pct = t["percent_complete"]
        total_dur += dur
        actual_w += dur * pct / 100.0
        if t["finish"] and t["finish"] <= cutoff_str:
            plan += 1

    actual = actual_w / total_dur * total if total_dur > 0 else 0.0
    target_pct = round(plan / total * 100) if total else 0
    actual_pct = round(actual / total * 100) if total else 0
    gap = target_pct - actual_pct

    return {
        "total": total,
        "plan": plan,
        "actual": round(actual, 2),
        "target_pct": target_pct,
        "actual_pct": actual_pct,
        "gap": gap,
    }


def calculate(file_paths, cutoff_dates=None):
    """Calculate Plan vs Actual for one or more MPP files at multiple dates.

    Args:
        file_paths: list of file paths (.mpp or .xml)
        cutoff_dates: list of date objects (default: [today])

    Returns:
        dict with 'dates', 'titles', 'workstreams' (nested by date)
    """
    if cutoff_dates is None:
        cutoff_dates = [date.today()]

    # Merge tasks from all files
    all_tasks = []
    titles = []
    for fp in file_paths:
        title, tasks = read_tasks(fp)
        titles.append(title)
        all_tasks.extend(tasks)

    ws_order, ws_tasks = _build_workstreams(all_tasks)

    # Calculate for each date
    results_by_date = []
    for d in cutoff_dates:
        cutoff_str = f"{d.isoformat()}T17:00:00"
        ws_results = []
        for ws_name in ws_order:
            wd = ws_tasks[ws_name]
            meta = wd["meta"]
            finish = meta["finish"]
            target = finish[:10].replace("-", "/") if finish else ""

            calc = _calc_ws(wd["leaves"], cutoff_str)

            # Check if MPP has native Plan%/Gap%
            native_plan = meta.get("planned_pct")
            native_gap = meta.get("gap_pct")
            pct = meta["percent_complete"]

            if native_plan is not None:
                ws_results.append({
                    "project": ws_name,
                    "target": target,
                    "milestone_task": calc["total"],
                    "plan": calc["plan"],
                    "actual": calc["actual"],
                    "target_pct": round(native_plan),
                    "actual_pct": round(pct),
                    "gap": round(native_gap) if native_gap is not None else round(native_plan - pct),
                    "source": "mpp",
                })
            else:
                ws_results.append({
                    "project": ws_name,
                    "target": target,
                    "milestone_task": calc["total"],
                    "plan": calc["plan"],
                    "actual": calc["actual"],
                    "target_pct": calc["target_pct"],
                    "actual_pct": calc["actual_pct"],
                    "gap": calc["gap"],
                    "source": "computed",
                })

        results_by_date.append({
            "date": d.isoformat(),
            "workstreams": ws_results,
        })

    return {
        "titles": titles,
        "dates": [d.isoformat() for d in cutoff_dates],
        "results": results_by_date,
    }


# ---------------------------------------------------------------------------
# Output formatters
# ---------------------------------------------------------------------------

def format_markdown(result):
    """Format result as Markdown tables (one per date)."""
    lines = []
    for entry in result["results"]:
        lines.append(f"### Plan vs Actual ({entry['date']})")
        lines.append("")
        lines.append("| Project | Target | Milestone Task | Plan | Actual | Target% | Actual% | Gap |")
        lines.append("|---------|--------|----------------|------|--------|---------|---------|-----|")

        for ws in entry["workstreams"]:
            gap_str = f"**{ws['gap']}%**" if abs(ws["gap"]) >= 5 else f"{ws['gap']}%"
            lines.append(
                f"| {ws['project']} | {ws['target']} | {ws['milestone_task']} "
                f"| {ws['plan']} | {ws['actual']} "
                f"| {ws['target_pct']}% | {ws['actual_pct']}% | {gap_str} |"
            )
        lines.append("")

    # Data source note
    sources = set()
    for entry in result["results"]:
        for ws in entry["workstreams"]:
            sources.add(ws.get("source", "computed"))
    if "mpp" in sources:
        lines.append("_Note: Plan% and Gap% sourced from MPP custom fields (Number4/Number3)._")
    else:
        lines.append("_Note: Plan% and Gap% computed from task finish dates and durations._")
    lines.append("")

    return "\n".join(lines)


def format_excel(result, output_path):
    """Write result to Excel with one sheet per date + a summary sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HEADER_FILL = PatternFill("solid", fgColor="4472C4")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    THIN_BORDER = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )
    RED_FILL = PatternFill("solid", fgColor="FFC7CE")
    GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
    YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")

    def _hdr(ws, row, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row, i, h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")

    def _cell(ws, row, col, value, fmt=None, fill=None, bold=False):
        c = ws.cell(row, col, value)
        c.border = THIN_BORDER
        if fmt:
            c.number_format = fmt
        if fill:
            c.fill = fill
        if bold:
            c.font = Font(bold=True)
        return c

    def _gap_fill(gap):
        if gap >= 10:
            return RED_FILL
        elif gap <= -5:
            return GREEN_FILL
        elif abs(gap) >= 5:
            return YELLOW_FILL
        return None

    wb = Workbook()

    # --- Summary sheet: all dates side by side ---
    ws_sum = wb.active
    ws_sum.title = "Summary"

    dates = [e["date"] for e in result["results"]]
    ws_names = [ws["project"] for ws in result["results"][0]["workstreams"]]

    # Header row 1: merged date headers
    row = 1
    _cell(ws_sum, row, 1, "Project", bold=True)
    _cell(ws_sum, row, 2, "Target", bold=True)
    _cell(ws_sum, row, 3, "Tasks", bold=True)
    col = 4
    for d in dates:
        for h in ["Plan", "Actual", "Target%", "Actual%", "Gap"]:
            c = _cell(ws_sum, row, col, f"{d} {h}" if h == "Plan" else h, bold=True)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            col += 1

    # Data rows
    for i, ws_name in enumerate(ws_names):
        row = i + 2
        first_ws = result["results"][0]["workstreams"][i]
        _cell(ws_sum, row, 1, ws_name, bold=True)
        _cell(ws_sum, row, 2, first_ws["target"])
        _cell(ws_sum, row, 3, first_ws["milestone_task"])

        col = 4
        for entry in result["results"]:
            ws = entry["workstreams"][i]
            _cell(ws_sum, row, col, ws["plan"])
            _cell(ws_sum, row, col + 1, ws["actual"])
            _cell(ws_sum, row, col + 2, ws["target_pct"] / 100, fmt="0%")
            _cell(ws_sum, row, col + 3, ws["actual_pct"] / 100, fmt="0%")
            _cell(ws_sum, row, col + 4, ws["gap"] / 100, fmt="0%",
                  fill=_gap_fill(ws["gap"]))
            col += 5

    # Auto-width
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 12
    ws_sum.column_dimensions["C"].width = 8
    for c_idx in range(4, col):
        from openpyxl.utils import get_column_letter
        ws_sum.column_dimensions[get_column_letter(c_idx)].width = 12

    # --- Per-date sheets ---
    for entry in result["results"]:
        d = entry["date"]
        sheet_name = d.replace("-", "")
        ws = wb.create_sheet(sheet_name)

        headers = ["Project", "Target", "Milestone Task", "Plan", "Actual",
                    "Target%", "Actual%", "Gap"]
        _hdr(ws, 1, headers)

        for r, wsdata in enumerate(entry["workstreams"], 2):
            _cell(ws, r, 1, wsdata["project"], bold=True)
            _cell(ws, r, 2, wsdata["target"])
            _cell(ws, r, 3, wsdata["milestone_task"])
            _cell(ws, r, 4, wsdata["plan"])
            _cell(ws, r, 5, wsdata["actual"])
            _cell(ws, r, 6, wsdata["target_pct"] / 100, fmt="0%")
            _cell(ws, r, 7, wsdata["actual_pct"] / 100, fmt="0%")
            _cell(ws, r, 8, wsdata["gap"] / 100, fmt="0%",
                  fill=_gap_fill(wsdata["gap"]))

        for col_letter, w in {"A": 30, "B": 12, "C": 14, "D": 8,
                               "E": 10, "F": 10, "G": 10, "H": 10}.items():
            ws.column_dimensions[col_letter].width = w

    output_dir = os.path.dirname(os.path.abspath(output_path))
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------

def _next_friday(d):
    """Return the Friday of the week containing d."""
    days_ahead = 4 - d.weekday()  # Friday = 4
    if days_ahead < 0:
        days_ahead += 7
    return d + timedelta(days=days_ahead)


def _week_dates(n_weeks, reference=None):
    """Generate n weekly cutoff dates (Fridays) starting from this week."""
    if reference is None:
        reference = date.today()
    friday = _next_friday(reference)
    return [friday + timedelta(weeks=i) for i in range(n_weeks)]


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Plan vs Actual gap analysis for MPP/MSPDI files")
    parser.add_argument("files", nargs="+", help="One or more MPP or XML files")
    parser.add_argument("--date", help="Single cutoff date YYYY-MM-DD")
    parser.add_argument("--dates", help="Comma-separated cutoff dates YYYY-MM-DD")
    parser.add_argument("--weeks", type=int, help="Generate N weekly cutoff dates (Fridays)")
    parser.add_argument("--json", action="store_true", help="Output JSON")
    parser.add_argument("--excel", help="Output Excel file path")
    args = parser.parse_args()

    # Determine cutoff dates
    if args.weeks:
        cutoff_dates = _week_dates(args.weeks)
    elif args.dates:
        cutoff_dates = [date.fromisoformat(d.strip()) for d in args.dates.split(",")]
    elif args.date:
        cutoff_dates = [date.fromisoformat(args.date)]
    else:
        cutoff_dates = [date.today()]

    # Validate files
    for f in args.files:
        if not os.path.isfile(f):
            print(f"Error: File not found: {f}", file=sys.stderr)
            sys.exit(2)

    result = calculate(args.files, cutoff_dates)

    if args.excel:
        path = format_excel(result, args.excel)
        print(f"Saved: {path}")
        print(f"Sources: {', '.join(result['titles'])}")
        print(f"Dates: {', '.join(result['dates'])}")
        print(f"Workstreams: {len(result['results'][0]['workstreams'])}")
        print(f"Sheets: Summary + {len(result['dates'])} date sheets")
    elif args.json:
        print(json.dumps(result, indent=2, ensure_ascii=False))
    else:
        print(format_markdown(result))


if __name__ == "__main__":
    main()
